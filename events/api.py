from django.http import JsonResponse, FileResponse
from timetable.models import Lesson, RecordsLesson
from register.models import User
from django.contrib.auth.decorators import login_required, permission_required
from json import loads
import openpyxl
from openpyxl.styles import Font


@login_required
@permission_required('events.view_event')
def exportData(response, pk):
    xltable = openpyxl.Workbook()
    lesson = Lesson.objects.get(pk=pk)
    records = RecordsLesson.objects.filter(lesson=pk).values_list('user', 'presence')

    sheet_name = lesson.name if lesson.name else lesson.theme

    xltable.create_sheet('list')
    active_sheet = xltable['list']

    active_sheet["A1"] = sheet_name
    active_sheet["A2"] = '№'
    active_sheet["B2"] = 'Статус'
    active_sheet['C2'] = 'Фамилия'
    active_sheet['D2'] = 'Имя'
    active_sheet['E2'] = 'Отчество'
    active_sheet['F2'] = 'Телефон'
    active_sheet['G2'] = 'Почта'
    active_sheet['H2'] = 'Место жительства'
    active_sheet['I2'] = 'Образовательное учреждение'
    active_sheet['J2'] = 'Класс/курс'


    for i in range(len(records)): 
        member_data = User.objects.filter(pk=records[i][0]).values_list('status', 'last_name', 'first_name', 'middle_name', 'phone', 'username', 'place', 'educational_institution', 'class_number')[0]

        active_sheet.cell(row=i+3, column=1).value = i+1
        if records[i][1]:
            active_sheet.cell(row=i+3, column=1).font = Font(color='5DE100')
        for c, data in enumerate(member_data):
            active_sheet.cell(row=i+3, column=c+2).value = data


    xltable.remove(xltable['Sheet'])
    xltable.save(f'export_data/event_{pk}.xlsx')

    file = open(f'export_data/event_{pk}.xlsx', "rb")
    response = FileResponse(file)
    
    return response


@login_required
@permission_required('events.view_event')
def allExportData(request):
    users = User.objects.all().order_by('-date_joined').values_list('id', 'status', 'last_name', 'first_name', 'middle_name', 'phone', 'username', 'place', 'educational_institution', 'class_number', named=True)

    lessons = Lesson.objects.all().order_by('lesson_date').values_list('id', 'name', named=True)


    xltable = openpyxl.Workbook()
    xltable.create_sheet('Участники')
    xltable.create_sheet('Посещаемость')
    users_sheet = xltable['Участники']
    presence_sheet = xltable['Посещаемость']

    users_sheet["A1"] = '№'
    users_sheet["B1"] = 'Статус'
    users_sheet['C1'] = 'Фамилия'
    users_sheet['D1'] = 'Имя'
    users_sheet['E1'] = 'Отчество'
    users_sheet['F1'] = 'Телефон'
    users_sheet['G1'] = 'Почта'
    users_sheet['H1'] = 'Место жительства'
    users_sheet['I1'] = 'Образовательное учреждение'
    users_sheet['J1'] = 'Класс/курс'

    presence_sheet["A1"] = '№'
    presence_sheet["B1"] = 'Фамилия'
    presence_sheet['C1'] = 'Имя'
    presence_sheet['D1'] = 'Всего'

    for i in range(len(lessons)):
        presence_sheet.cell(row=1, column=i+5).value = lessons[i][1]

    for i in range(len(users)):
        user_id = users[i][0] 
        user_data = users[i][1:]

        users_sheet.cell(row=i+2, column=1).value = i+1
        presence_sheet.cell(row=i+2, column=1).value = i+1
        presence_sheet.cell(row=i+2, column=2).value = user_data[3]
        presence_sheet.cell(row=i+2, column=3).value = user_data[2]
        presence_sheet.cell(row=i+2, column=4).value = len(RecordsLesson.objects.filter(user=user_id, presence=True))

        for c, data in enumerate(user_data):
            users_sheet.cell(row=i+2, column=c+2).value = data

        for c in range(len(lessons)):
            value = '+' if RecordsLesson.getPresence(user_id, lessons[c][0]) else '-'
            presence_sheet.cell(row=i+2, column=c+5).value = value

            
    xltable.remove(xltable['Sheet'])
    xltable.save(f'export_data/users.xlsx')

    file = open('export_data/users.xlsx', "rb")
    response = FileResponse(file)
    
    return response


@login_required
@permission_required('events.view_event')
def markVisit(request, pk):
    if request.method == 'POST':
        user_email = loads(request.body)['user_email']
        user = User.objects.get(email=user_email)
        lesson = Lesson.objects.get(pk=pk)

        record = RecordsLesson.objects.get(lesson=lesson.id, user=user.id)

        if record.presence:
            return JsonResponse({'error': 'Был отмечен ранее'})

        record.presence = True
        record.save()
        user.score += lesson.points
        user.save()
        return JsonResponse({'result': 'Успешно отмечен'})


@login_required
@permission_required('events.view_event')
def cancelVisit(request, pk):
    if request.method == 'POST':
        user_email = loads(request.body)['user_email']
        user = User.objects.get(email=user_email)
        lesson = Lesson.objects.get(pk=pk)

        record = RecordsLesson.objects.get(lesson=lesson.id, user=user.id)

        if not record.presence:
            return JsonResponse({'error': 'Ошибка'})

        record.presence = False
        record.save()
        user.score -= lesson.points
        user.save()
        return JsonResponse({'result': 'Успешно'})